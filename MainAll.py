import math as m
import cv2
import numpy as np
import mediapipe as mp
import matplotlib.pyplot as plt
import time
import pandas as pd
from openpyxl import load_workbook



#Weight = input("What is weight of object in KG ?")
# Time
time_ = 0

# Initializing mediapipe pose class.
# mp_pose = mp.solutions.pose
mp_pose = mp.solutions.holistic

mp_drawing_styles = mp.solutions.drawing_styles
# mp_holistic = mp.solutions.holistic

# Setting up the Pose function.
# pose = mp_pose.Pose(static_image_mode=True, min_detection_confidence=0.3, model_complexity=2)

# Initializing mediapipe drawing class, useful for annotation.
mp_drawing = mp.solutions.drawing_utils 

def extract_keypoints(results):
    pose = np.array([[res.x, res.y, res.z, res.visibility] for res in results.pose_landmarks.landmark]).flatten() if results.pose_landmarks else np.zeros(132)
    face = np.array([[res.x, res.y, res.z] for res in results.face_landmarks.landmark]).flatten() if results.face_landmarks else np.zeros(1404)
    left_hand = np.array([[res.x, res.y, res.z] for res in results.left_hand_landmarks.landmark]).flatten() if results.left_hand_landmarks else np.zeros(63)
    right_hand = np.array([[res.x, res.y, res.z] for res in results.right_hand_landmarks.landmark]).flatten() if results.right_hand_landmarks else np.zeros(63)
    return np.concatenate([pose, face, left_hand, right_hand])

def extract_keypoints_Pose(results):
        pose = np.array([[res.x, res.y, res.z, res.visibility] for res in results.pose_landmarks.landmark]).flatten() if results.pose_landmarks else np.zeros(33*4)
        return np.concatenate([pose])
    
def extract_keypoints_Pose_(results):
        pose = []
        for i in range(len(results.pose_landmarks.landmark)):
            sample = [results.pose_landmarks.landmark[i].x, results.pose_landmarks.landmark[i].y]
            pose.append(sample)
        pose = np.array(pose)
        pose = pose.reshape((1,*pose.shape))
        return pose

def detectPose(image, pose, display=True):
    '''
    This function performs pose detection on an image.
    Args:
        image: The input image with a prominent person whose pose landmarks needs to be detected.
        pose: The pose setup function required to perform the pose detection.
        display: A boolean value that is if set to true the function displays the original input image, the resultant image, 
                 and the pose landmarks in 3D plot and returns nothing.
    Returns:
        output_image: The input image with the detected pose landmarks drawn.
        landmarks: A list of detected landmarks converted into their original scale.
    '''
    
    # Create a copy of the input image.
    output_image = image.copy()
    
    # Convert the image from BGR into RGB format.
    imageRGB = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    # Perform the Pose Detection.
    results = pose.process(imageRGB)
    
    # Retrieve the height and width of the input image.
    height, width, _ = image.shape
    
    # Initialize a list to store the detected landmarks.
    landmarks = []
    
    # keypoint = extract_keypoints_Pose_(results)
    # print(keypoint.shape)
    
    
    
    # Check if any landmarks are detected.
    if results.pose_landmarks:
    
        # Draw Pose landmarks on the output image.
        mp_drawing.draw_landmarks(image=output_image, landmark_list=results.pose_landmarks,connections=mp_pose.POSE_CONNECTIONS,landmark_drawing_spec=mp_drawing_styles.get_default_pose_landmarks_style())
        mp_drawing.draw_landmarks(image=output_image,landmark_list=results.left_hand_landmarks,connections=mp_pose.HAND_CONNECTIONS,landmark_drawing_spec=mp_drawing_styles.get_default_hand_landmarks_style())
        mp_drawing.draw_landmarks(image=output_image, landmark_list=results.right_hand_landmarks, connections=mp_pose.HAND_CONNECTIONS,landmark_drawing_spec=mp_drawing_styles.get_default_hand_landmarks_style())
        
        # mp_drawing.draw_landmarks(image=output_image, landmark_list=results.face_landmarks,connections=mp_pose.FACEMESH_CONTOURS,landmark_drawing_spec=None,connection_drawing_spec=mp_drawing_styles.get_default_face_mesh_contours_style())
        #mp_drawing.draw_landmarks(image=output_image, landmark_list=results.face_landmarks,connections=mp_pose.FACEMESH_TESSELATION,landmark_drawing_spec=None,connection_drawing_spec=mp_drawing_styles.get_default_face_mesh_tesselation_style())
        
        # Iterate over the detected landmarks.
        for landmark in results.pose_landmarks.landmark:
            # Append the landmark into the list.
            landmarks.append((int(landmark.x * width), int(landmark.y * height),
                                  (landmark.z * width)))
            
        # Append the landmark LEFT Hand into the list. 
        if results.right_hand_landmarks:
            for landmark in results.right_hand_landmarks.landmark:
                # Append the landmark into the list.
                landmarks.append((int(landmark.x * width), int(landmark.y * height),
                                    (landmark.z * width)))
        else:
             for i in range(21):
                # Append the landmark into the list.
                landmarks.append((int(0* width), int(0 * height),
                                    (0 * width)))
        
        # Append the landmark RIGHT Hand into the list. 
        if results.left_hand_landmarks:
            for landmark in results.left_hand_landmarks.landmark:
                # Append the landmark into the list.
                landmarks.append((int(landmark.x * width), int(landmark.y * height),
                                    (landmark.z * width)))
            # print(len(results.left_hand_landmarks.landmark))
        else:
             for i in range(21):
                # Append the landmark into the list.
                landmarks.append((int(0* width), int(0 * height),
                                    (0 * width)))
        
        # print(len(landmarks))
    
    # Check if the original input image and the resultant image are specified to be displayed.
    if display:
    
        # Display the original input image and the resultant image.
        plt.figure(figsize=[22,22])
        plt.subplot(121);plt.imshow(image[:,:,::-1]);plt.title("Original Image");plt.axis('off');
        plt.subplot(122);plt.imshow(output_image[:,:,::-1]);plt.title("Output Image");plt.axis('off');
        
        # Also Plot the Pose landmarks in 3D.
        mp_drawing.plot_landmarks(results.pose_world_landmarks, mp_pose.POSE_CONNECTIONS)
        
    # Otherwise
    else:
        
        # Return the output image and the found landmarks.
        return output_image, landmarks
    
def calculateAngle(landmark1, landmark2, landmark3):
    '''
    This function calculates angle between three different landmarks.
    Args:
        landmark1: The first landmark containing the x,y and z coordinates.
        landmark2: The second landmark containing the x,y and z coordinates.
        landmark3: The third landmark containing the x,y and z coordinates.
    Returns:
        angle: The calculated angle between the three landmarks.

    '''

    # Get the required landmarks coordinates.
    x1, y1, _ = landmark1
    x2, y2, _ = landmark2
    x3, y3, _ = landmark3

    # Calculate the angle between the three points
    angle = m.degrees(m.atan2(y3 - y2, x3 - x2) - m.atan2(y1 - y2, x1 - x2))
    
    # Check if the angle is less than zero.
    if angle < 0:

        # Add 360 to the found angle.
        angle += 360
    
    # Return the calculated angle.
    return angle

def calculateDistance(landmark1, landmark2):
    x1, y1, _ = landmark1
    x2, y2, _ = landmark2

    # Calculate the Distance between the two points
    dis = m.sqrt( ((x2 - x1)**2)+((y2 - y1))**2)

    
    # Return the calculated Distance.
    return dis

# Calculate distance
def findDistance(x1, y1, x2, y2):
    dist = m.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
    return dist

# Calculate angle.
def findAngle(x1, y1, x2, y2):
    theta = m.acos((y2 - y1) * (-y1) / (m.sqrt(
        (x2 - x1) ** 2 + (y2 - y1) ** 2) * y1))
    degree = int(180 / m.pi) * theta
    return degree

left_elbow_angle = 0;
right_elbow_angle = 0;
left_shoulder_angle = 0;
right_shoulder_angle = 0;
left_wrist_angle = 0;
right_wrist_angle = 0;
middle_knee_angle = 0;
trunk_angle = 0;
neck_angle = 0;

left_elbow_angle_previous = 0;
right_elbow_angle_previous = 0;
left_shoulder_angle_previous = 0;
right_shoulder_angle_previous = 0;
left_wrist_angle_previous = 0;
right_wrist_angle_previous = 0;
middle_knee_angle_previous = 0;
trunk_angle_previous = 0;
neck_angle_previous = 0;

left_elbow_angle_diff = 0;
right_elbow_angle_diff = 0;
left_shoulder_angle_diff = 0;
right_shoulder_angle_diff = 0;
left_wrist_angle_diff = 0;
right_wrist_angle_diff = 0;
middle_knee_angle_diff = 0;
trunk_angle_diff = 0;
neck_angle_diff = 0;

Angle_previous = []

def updateAngle_previous():
    # Angle_previous = []
    left_elbow_angle_previous = left_elbow_angle;
    right_elbow_angle_previous = right_elbow_angle;
    left_shoulder_angle_previous = left_shoulder_angle;
    right_shoulder_angle_previous = right_shoulder_angle;
    left_wrist_angle_previous = left_wrist_angle;
    right_wrist_angle_previous = right_wrist_angle;
    middle_knee_angle_previous = middle_knee_angle;
    trunk_angle_previous = trunk_angle;
    neck_angle_previous = neck_angle;
    Angle_previous.append([left_elbow_angle_previous,right_elbow_angle_previous,left_shoulder_angle_previous,right_shoulder_angle_previous,left_wrist_angle_previous,right_wrist_angle_previous, middle_knee_angle_previous, trunk_angle_previous, neck_angle_previous])


Angle_diff = []


#Kid's
def find_rula(input1,input2,input3,input4, excel_file):
    x = str(input3) + str(input4)
    # print(x)

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Search for a matching row in the Excel table
    for row in sheet.iter_rows(min_row=5, values_only=True):
            if (input1,input2)==row[:2]:
                value= tuple(row[2:])
                key=("11", "12", "21", "22", "31", "32", "41", "42", "51", "52", "61", "62", "71", "72", "13", "14", "15", "16", "17", "18", "19")
                dict={k:v for (k,v) in zip(key, value)}
                if x in key:
                    return(str(dict[x]))
            elif input3 == 1:
                value = tuple(row[1:])
                key = ("11", "12", "21", "22", "31", "32", "41", "42", "51", "52", "61", "62", "71", "72", "13", "14", "15", "16", "17", "18", "19")
                dict = {k: v for (k, v) in zip(key, value)}
                if x in key:
                    return str(dict[x])

    workbook.close()

def updateAngle_diff():
    # Angle_diff = []
    left_elbow_angle_diff = left_elbow_angle - left_elbow_angle_previous;
    right_elbow_angle_diff = right_elbow_angle - right_elbow_angle_previous;
    left_shoulder_angle_diff = left_shoulder_angle - left_shoulder_angle_previous;
    right_shoulder_angle_diff = right_shoulder_angle - right_shoulder_angle_previous;
    left_wrist_angle_diff = left_wrist_angle - left_wrist_angle_previous;
    right_wrist_angle_diff = right_wrist_angle - right_wrist_angle_previous;
    middle_knee_angle_diff = middle_knee_angle - middle_knee_angle_previous;
    trunk_angle_diff = trunk_angle - trunk_angle_previous;
    neck_angle_diff = neck_angle - neck_angle_previous
    Angle_diff.append([left_elbow_angle_diff,right_elbow_angle_diff,left_shoulder_angle_diff,right_shoulder_angle_diff,left_wrist_angle_diff,right_wrist_angle_diff, middle_knee_angle_diff, trunk_angle_diff, neck_angle_diff])



def classifyPose(landmarks, output_image, display=False):
    global time_
    global Angle_previous
    global Angle_diff
    
    global left_elbow_angle
    global right_elbow_angle
    global left_shoulder_angle
    global right_shoulder_angle
    global left_wrist_range_angle
    global right_wrist_range_angle
    global left_wrist_angle
    global right_wrist_angle
    global middle_knee_angle
    global trunk_angle
    global neck_angle

    global left_elbow_angle_previous
    global right_elbow_angle_previous
    global left_shoulder_angle_previous
    global right_shoulder_angle_previous
    global left_wrist_angle_previous
    global right_wrist_angle_previous
    global middle_knee_angle_previous
    global trunk_angle_previous
    global neck_angle_previous

    global left_elbow_angle_diff
    global right_elbow_angle_diff
    global left_shoulder_angle_diff
    global right_shoulder_angle_diff
    global left_wrist_angle_diff
    global right_wrist_angle_diff
    global middle_knee_angle_diff
    global trunk_angle_diff
    global neck_angle_diff        

    global left_knee_score
    global Trunk_score
    global Neck_score


    '''
    This function classifies yoga poses depending upon the angles of various body joints.
    Args:
        landmarks: A list of detected landmarks of the person whose pose needs to be classified.
        output_image: A image of the person with the detected pose landmarks drawn.
        display: A boolean value that is if set to true the function displays the resultant image with the pose label 
        written on it and returns nothing.
    Returns:
        output_image: The image with the detected pose landmarks drawn and pose label written.
        label: The classified pose label of the person in the output_image.

    '''
    
    # Initialize the label of the pose. It is not known at this stage.
    label = 'Unknown Pose'

    # Specify the color (Red) with which the label will be written on the image.
    color = (0, 0, 255)
    
    # Calculate the required angles.
    #----------------------------------------------------------------------------------------------------------------
    
    # Get the angle between the left shoulder, elbow and wrist points. 
    right_elbow_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value],
                                      landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value],
                                      landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value])
    
    # Calculate the angle between the left elbow, left wrist, and midpoint between left index and left pinky
    left_wrist_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value],
                                  landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value],
                                  landmarks[mp_pose.PoseLandmark.LEFT_THUMB.value])
    
    # Calculate the angle between the left elbow, left wrist, and midpoint between left index and left pinky
    right_wrist_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value],
                                  landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value],
                                  landmarks[mp_pose.PoseLandmark.RIGHT_THUMB.value])

    
    # Get the angle between the right shoulder, elbow and wrist points. 
    left_elbow_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value],
                                       landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value],
                                       landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value])   
    
    # Get the angle between the left elbow, shoulder and hip points. 
    right_shoulder_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value],
                                         landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value],
                                         landmarks[mp_pose.PoseLandmark.LEFT_HIP.value])

    # Get the angle between the right hip, shoulder and elbow points. 
    left_shoulder_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value],
                                          landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value],
                                          landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value])

    # Get the angle between the left hip, knee and ankle points. 
    right_knee_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.LEFT_HIP.value],
                                     landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value],
                                     landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value])

    # Get the angle between the right hip, knee and ankle points 
    left_knee_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value],
                                      landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value],
                                      landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value])
    
    middle_knee_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value],
                                      landmarks[int((mp_pose.PoseLandmark.RIGHT_HIP.value - mp_pose.PoseLandmark.LEFT_HIP.value)/2)],
                                      landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value])
    


    trunk_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value],
                                      landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value],
                                      landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value])

    l_shldr_x = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value][0]
    l_shldr_y = landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value][1]

    r_shldr_x = landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value][0]
    r_shldr_y = landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value][1]

    l_ear_x = landmarks[mp_pose.PoseLandmark.LEFT_EAR.value][0]
    l_ear_y = landmarks[mp_pose.PoseLandmark.LEFT_EAR.value][1]

    l_hip_x = landmarks[mp_pose.PoseLandmark.LEFT_HIP.value][0]
    l_hip_y = landmarks[mp_pose.PoseLandmark.LEFT_HIP.value][1]

    neck_inclination = int(findAngle(l_shldr_x, l_shldr_y, l_ear_x, l_ear_y) - 35)
    torso_inclination = int(findAngle(l_hip_x, l_hip_y, l_shldr_x, l_shldr_y) - 6)

    LEFT_HAND = 33
    RIGHT_HAND = 33+21
    
    # Range of Wrist
    #Custom
   

     
    #HAND+POSE
    left_wrist_range_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value],
                                      landmarks[mp_pose.HandLandmark.WRIST.value+LEFT_HAND],
                                      landmarks[mp_pose.PoseLandmark.LEFT_INDEX.value+LEFT_HAND])
    
    right_wrist_range_angle = calculateAngle(landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value],
                                      landmarks[mp_pose.HandLandmark.WRIST.value+RIGHT_HAND],
                                      landmarks[mp_pose.PoseLandmark.RIGHT_INDEX.value+RIGHT_HAND])  
    
    
    # Get the angle between the wrist, thumb_tip and pinky_tip
    left_wrist_angle = calculateAngle(landmarks[mp_pose.HandLandmark.THUMB_TIP.value+LEFT_HAND],
                                      landmarks[mp_pose.HandLandmark.WRIST.value+LEFT_HAND],
                                      landmarks[mp_pose.HandLandmark.PINKY_TIP.value+LEFT_HAND]) 

    right_wrist_angle = calculateAngle(landmarks[mp_pose.HandLandmark.THUMB_TIP.value+RIGHT_HAND],
                                      landmarks[mp_pose.HandLandmark.WRIST.value+RIGHT_HAND],
                                      landmarks[mp_pose.HandLandmark.PINKY_TIP.value+RIGHT_HAND]) 
    


    Neck_score = 1
    if 0 < neck_inclination <= 10 :
        Neck_score = 1
         
    elif    10 < neck_inclination <= 20 :  
        Neck_score = 2
        
    elif    20 < neck_inclination :
        Neck_score = 3
        
    elif  neck_inclination < 0 :
        Neck_score = 4

    Trunk_score = 1
    if 0 < torso_inclination <= 10 :
        Trunk_score = 1
         
    elif    10 < torso_inclination <= 20 :  
        Trunk_score = 2
        
    elif    20 < torso_inclination <= 60 :
        Trunk_score = 3
        
    elif  60 <torso_inclination  :
        Trunk_score = 4
       
    text_posx = 20;
    text_step = 40;

    left_knee_score = 1
    right_knee_score = 1
    if 360 < middle_knee_angle >= 345:
        if 0 <= abs(left_knee_angle - right_knee_angle) <= 10:
            left_knee_score = 1
            right_knee_score = 1
        else:
            left_knee_score = 2
            right_knee_score = 2
    elif -30 < middle_knee_angle < 345:
        left_knee_score = 2
        right_knee_score = 2
    

    left_thumb_index_distance = calculateDistance(landmarks[mp_pose.HandLandmark.THUMB_TIP.value+LEFT_HAND],
                                      landmarks[mp_pose.HandLandmark.INDEX_FINGER_TIP.value+LEFT_HAND])

    right_thumb_index_distance = calculateDistance(landmarks[mp_pose.HandLandmark.THUMB_TIP.value+RIGHT_HAND],
                                      landmarks[mp_pose.HandLandmark.INDEX_FINGER_TIP.value+RIGHT_HAND])
    
    left_thumb_pinky_distance = calculateDistance(landmarks[mp_pose.HandLandmark.THUMB_TIP.value+LEFT_HAND],
                                      landmarks[mp_pose.HandLandmark.PINKY_TIP.value+LEFT_HAND])

    right_thumb_pinky_distance = calculateDistance(landmarks[mp_pose.HandLandmark.THUMB_TIP.value+RIGHT_HAND],
                                      landmarks[mp_pose.HandLandmark.PINKY_TIP.value+RIGHT_HAND])
    #---
    
    
    # Show Circlr Distance Point
    cirlcr1_posx ,cirlcr1_posy,_= landmarks[mp_pose.HandLandmark.THUMB_TIP.value+LEFT_HAND]
    cirlcr2_posx ,cirlcr2_posy,_= landmarks[mp_pose.HandLandmark.INDEX_FINGER_TIP.value+LEFT_HAND]
    circle_center = (int((cirlcr1_posx+cirlcr2_posx)*0.5),int((cirlcr1_posy+cirlcr2_posy)*0.5))
    cv2.circle(output_image,circle_center,int(left_thumb_index_distance/2.5),(0,0,0),-1)
    
    cirlcr1_posx ,cirlcr1_posy,_= landmarks[mp_pose.HandLandmark.THUMB_TIP.value+RIGHT_HAND]
    cirlcr2_posx ,cirlcr2_posy,_= landmarks[mp_pose.HandLandmark.INDEX_FINGER_TIP.value+RIGHT_HAND]
    circle_center = (int((cirlcr1_posx+cirlcr2_posx)*0.5),int((cirlcr1_posy+cirlcr2_posy)*0.5))
    cv2.circle(output_image,circle_center,int(right_thumb_index_distance/2.5),(0,0,0),-1)

    # Draw Data
    
    text_posx = 20;
    text_step = 40
    # Left upper arm
    #cv2.putText(output_image, "L_upper_angle : " + str("{:0.2f}".format(left_shoulder_angle)), (10, text_posx),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    
    L_upper_arm_score = 1
    if 90 < left_shoulder_angle <= 180:
        L_upper_arm_score=4
        
    elif left_shoulder_angle<=20:
        L_upper_arm_score=1
        
    elif left_shoulder_angle<=45:
        L_upper_arm_score=2
        
    elif left_shoulder_angle <=90:
        L_upper_arm_score=3

    cv2.putText(output_image, "L_upper_arm_score : " + str("{:0.2f}".format(L_upper_arm_score)), (500, text_posx),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)    
    
    
    # Right upper arm
    
    #cv2.putText(output_image, "R_upper_angle : " + str("{:0.2f}".format(right_shoulder_angle)), (10, text_posx+text_step*1),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    R_upper_arm_score = 1
    if 90 <right_shoulder_angle <=180:
        R_upper_arm_score=4
        
    elif right_shoulder_angle<=20:
        R_upper_arm_score=1
        
    elif right_shoulder_angle<=45:
        R_upper_arm_score=2
        
    elif right_shoulder_angle <=90:
        R_upper_arm_score=3
    
    cv2.putText(output_image, "R_upper_arm_score : " + str("{:0.2f}".format(R_upper_arm_score)), (500, text_posx+text_step*1),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)
    
    # Left_Lower_arm เสร็จ
    #cv2.putText(output_image, "L_elbow_angle : " + str("{:0.2f}".format(left_elbow_angle)), (10, text_posx+text_step*2),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    L_lower_arm_score = 1
    if  left_elbow_angle < 240:
        L_lower_arm_score=2
        
    elif 240<= left_elbow_angle<280:
        L_lower_arm_score=1
        
    elif left_elbow_angle >=280:
        L_lower_arm_score=2
        
    cv2.putText(output_image, "L_lower_arm_score : " + str("{:0.2f}".format(L_lower_arm_score)), (500, text_posx+text_step*2),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)

    # Right_Lower_arm เสร็จ
   # cv2.putText(output_image, "R_elbow_angle : " + str("{:0.2f}".format(right_elbow_angle)), (10, text_posx+text_step*3),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    R_lower_arm_score = 1
    if  right_elbow_angle < 240:
        R_lower_arm_score=2
        
    elif 240<= right_elbow_angle<280:
        R_lower_arm_score=1
        
    elif right_elbow_angle>=280:
        R_lower_arm_score=2
    cv2.putText(output_image, "R_lower_arm_score : " + str("{:0.2f}".format(R_lower_arm_score)), (500, text_posx+text_step*3),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2) 
    
    
    
    
   # Left_wrist_Range
    #cv2.putText(output_image, "L_Wrist_Range_Angle : " + str("{:0.2f}".format(left_wrist_range_angle)), (10, text_posx+text_step*4),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    L_wrist_range_score = 1 
    if  290 < left_wrist_range_angle < 320:
    
        L_wrist_range_score = 1 
    
    elif 260 <= left_wrist_range_angle < 290:
        L_wrist_range_score=2
    elif 200 <= left_wrist_range_angle < 260:
        L_wrist_range_score=3

    elif 320 <=left_wrist_range_angle < 360:
        L_wrist_range_score=2
    elif 0 < left_wrist_range_angle < 100:
        L_wrist_range_score=3
    cv2.putText(output_image, "L_wristrange_score : " + str("{:0.2f}".format(L_wrist_range_score)), (500, text_posx+text_step*4),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)



    # Right_wrist_Range
    #cv2.putText(output_image, "R_Wrist_Range_Angle : " + str("{:0.2f}".format(right_wrist_range_angle)), (10, text_posx+text_step*5),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    R_wrist_range_score = 1
    if  30 < right_wrist_range_angle < 60:
        
        R_wrist_range_score=1
    
    elif 60 <= right_wrist_range_angle < 90:
        R_wrist_range_score=2
    elif 90 <= right_wrist_range_angle < 180:
        R_wrist_range_score=3

    elif 0 <right_wrist_range_angle < 30:
        R_wrist_range_score=2
    elif 280 < right_wrist_range_angle < 360:
        R_wrist_range_score=3
    cv2.putText(output_image, "R_wristrange_score : " + str("{:0.2f}".format(R_wrist_range_score)), (500, text_posx+text_step*5),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)
    
    
    #Left_wrist_twist
   # cv2.putText(output_image, "L_wrist_angle : " + str("{:0.2f}".format(left_wrist_angle)), (10, text_posx+text_step*6),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    L_wrist_range_score = 1
    if   left_wrist_angle < 100:
        L_wrist_twist_score=2
        
    elif 100<= left_wrist_angle <360:
        L_wrist_twist_score=1  
    cv2.putText(output_image, "L_wrist_twist_score : " + str("{:0.2f}".format(L_wrist_twist_score)), (500, text_posx+text_step*6),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2) 
   
    #Right_wrist_twist
   # cv2.putText(output_image, "R_wrist_angle : " + str("{:0.2f}".format(right_wrist_angle)), (10, text_posx+text_step*7),cv2.FONT_HERSHEY_PLAIN, 1.3, (0,0,255), 2)
    if   right_wrist_angle < 100:
        R_wrist_twist_score=1
        
    elif 100<= right_wrist_angle <360:
        R_wrist_twist_score=2  
    cv2.putText(output_image, "R_wrist_twist_score : " + str("{:0.2f}".format(R_wrist_twist_score)), (500, text_posx+text_step*7),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2) 
 #---------------------------------------------------------------------------------------------------------------------
    #FOR LEFT TABLE A
    LA = str(find_rula(L_upper_arm_score,L_lower_arm_score,L_wrist_range_score,L_wrist_twist_score,'TABLE_A.xlsx'))
    cv2.putText(output_image,"Left Table A : " + LA, (10, text_posx),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)
    #print(find_rula(L_upper_arm_score,L_upper_arm_score,L_wrist_range_score,L_wrist_twist_score))

    

    #FOR RIGHT TABLE A
    RA = str(find_rula(R_upper_arm_score,R_lower_arm_score,R_wrist_range_score,R_wrist_twist_score,'TABLE_A.xlsx'))
    cv2.putText(output_image,"Right Table A : " + RA, (10, text_posx+text_step*1),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)
    # print(find_rula(R_upper_arm_score,L_lower_arm_score,L_wrist_range_score,L_wrist_twist_score))

    #Foe Table B
    #LB = str(find_rula(Neck_score,Trunk_score,left_knee_score,'TABLE_B.xlsx'))
    #cv2.putText(output_image,"Table B : " + LB, (10, text_posx+text_step*2),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)


    
    # middle_knee_angle = 360
    left_knee_score = 1
    right_knee_score = 1

    if 360 < middle_knee_angle >= 345:
        if abs(left_knee_angle - right_knee_angle) <= 10 & abs(left_knee_angle - right_knee_angle) >= 0:
            # print(f"The absolute difference between {x} and {y} is within ±10.")
            left_knee_score = 1
            right_knee_score = 1
        else:
            # print(f"The absolute difference between {x} and {y} is greater than ±10.")
            left_knee_score = 2
            right_knee_score = 2
    elif 0 < middle_knee_angle < 345:
        left_knee_score = 2
        right_knee_score = 2
        
    LB = str(find_rula(1, Neck_score, Trunk_score, left_knee_score, 'TABLE_B.xlsx'))
    RB = LB
    cv2.putText(output_image,"Left&Right Table B : " + LB, (10, text_posx+text_step*2),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)

    LC = str(find_rula(1, LA, 1, LB, 'TABLE_C.xlsx'))
    cv2.putText(output_image,"Left Table C : " + LC, (10, text_posx+text_step*3),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)
    RC = str(find_rula(1, RA, 1, RB, 'TABLE_C.xlsx'))
    cv2.putText(output_image,"Right Table C : " + RC, (10, text_posx+text_step*4),cv2.FONT_HERSHEY_PLAIN, 1.3, (127, 20, 0), 2)


    # print(left_knee_score)
    # print(right_knee_score)
    # print(middle_knee_angle)
  #--------------------------------------------






# TABLE A LEFT
# input1 = int(input("Enter upper_arm_score_Left : "))
# input2 = int(input("Enter lower_arm_score_Left  : "))
# input3 = int(input("Enter wrist_score_Left  : "))
# input4 = int(input("Enter wrist_twist_score_Left  : "))



    
    
    #----------------------------------------------------------------------------------------------------------------
    
    # Update Diff Angle
    if(time.time() - time_ > 0.2):
        
        left_elbow_angle_diff = abs(left_elbow_angle - left_elbow_angle_previous);
        right_elbow_angle_diff = abs(right_elbow_angle - right_elbow_angle_previous);
        left_shoulder_angle_diff = abs(left_shoulder_angle - left_shoulder_angle_previous);
        right_shoulder_angle_diff = abs(right_shoulder_angle - right_shoulder_angle_previous);
        left_wrist_angle_diff = abs(left_wrist_angle - left_wrist_angle_previous);
        right_wrist_angle_diff = abs(right_wrist_angle - right_wrist_angle_previous);
        Angle_diff.append([left_elbow_angle_diff,right_elbow_angle_diff,left_shoulder_angle_diff,right_shoulder_angle_diff,left_wrist_angle_diff,right_wrist_angle_diff])

        left_elbow_angle_previous = left_elbow_angle;
        right_elbow_angle_previous = right_elbow_angle;
        left_shoulder_angle_previous = left_shoulder_angle;
        right_shoulder_angle_previous = right_shoulder_angle;
        left_wrist_angle_previous = left_wrist_angle;
        right_wrist_angle_previous = right_wrist_angle;
        
        Angle_previous = []
        Angle_diff = []
        time_ = time.time()
    
    if display:
    
        # Display the resultant image.
        plt.figure(figsize=[10,10])
        plt.title("Output Image");plt.axis('off');plt.imshow(output_image[:,:,::-1]);plt.show()
        
    else:
        
        # Return the output image and the classified label.
        return output_image, label
