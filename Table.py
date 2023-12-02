from openpyxl import load_workbook

def find_rula(input1,input2,input3,input4, excel_file):
    x = str(input3) + str(input4)
    # print(x)

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Search for a matching row in the Excel table
    for row in sheet.iter_rows(min_row=5, values_only=True):
            if (input1,input2)==row[:2]:
                value= tuple(row[2:])
                key=("11","12","21","22","31","32","41","42")
                dict={k:v for (k,v) in zip(key, value)}
                if x in key:
                    return(str(dict[x]))
            elif input2 == 'null' and input1 == row[0]:
                value = tuple(row[1:])
                key = ("11", "12", "21", "22", "31", "32", "41", "42")
                dictionary = {k: v for (k, v) in zip(key, value)}
                if x in key:
                    return str(dictionary[x])

    workbook.close()